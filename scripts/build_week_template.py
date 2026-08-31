"""
Builds docs/templates/hooslog_week_plan_template.xlsx — the coach's weekly
upload form, reworked for TWO training groups (distance + mid-distance).

Why this script exists: the template IS the importer's format contract
(CLAUDE.md locked 23). A hand-edited binary drifts from the parser and nobody
can see how it was made. Regenerate with:  python3 scripts/build_week_template.py

What changed from v1:
  Week Plan!C  WORKOUT PLAN  ->  DISTANCE PLAN
  Week Plan!D  (new)             MID-DISTANCE PLAN
  Goals!D      (new)             GROUP  (dropdown: Distance / Mid-D; blank = no change)
  Goals rows   pre-filled with the 30 rostered athletes (no example row to delete)
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY   = "FF232D4B"
GREY   = "FF5C6478"
YELLOW = "FFFFFF00"
WHITE  = "FFFFFFFF"

title_f  = Font(name="Arial", size=16, bold=True, color=NAVY)
head_f   = Font(name="Arial", size=10, bold=True, color=WHITE)
label_f  = Font(name="Arial", size=10, bold=True, color=NAVY)
body_f   = Font(name="Arial", size=10)
hint_f   = Font(name="Arial", size=9, color=GREY)
navy_fill   = PatternFill("solid", fgColor=NAVY)
yellow_fill = PatternFill("solid", fgColor=YELLOW)
thin = Side(style="thin", color="FFBFC5D2")
bb   = Border(bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")

# Roster: supabase/migrations/0005_closed_roster.sql (30 athletes), by last name.
ROSTER = [
    ("Henry Elijah Acorn", "dms7jc@virginia.edu"),
    ("Adam Christopher Balewicz", "hww4nw@virginia.edu"),
    ("Henry Trepagnier Birge", "qjp4nm@virginia.edu"),
    ("Ciaran Donnacha Brosnan", "wzz9ed@virginia.edu"),
    ("Shane Makana Brosnan", "atu4xv@virginia.edu"),
    ("Aidan Timothy Cox", "sen4zu@virginia.edu"),
    ("Philip David Cupial", "hju5az@virginia.edu"),
    ("Trent W Daniels", "gyu5nm@virginia.edu"),
    ("James Joseph Donahue", "rwd8an@virginia.edu"),
    ("Cayden Wayne Dyer", "kma8am@virginia.edu"),
    ("Tyler William Edson", "ret8ve@virginia.edu"),
    ("Quinn Dulany Eliason", "dtw3fe@virginia.edu"),
    ("James Bruce Ford", "jbe9ns@virginia.edu"),
    ("Ben Isaac Godish", "qnp3nj@virginia.edu"),
    ("Sean Warren Gray", "euu8xk@virginia.edu"),
    ("Cooper Davis Groat", "cgd2va@virginia.edu"),
    ("Luke William Hnatt", "tsv8cu@virginia.edu"),
    ("Andrew Graham Jones", "hub9fh@virginia.edu"),
    ("Alex Leath", "kxy2qc@virginia.edu"),
    ("Kayden Thomas Lightner", "nvv5qk@virginia.edu"),
    ("Brenden Michael McMahon", "zwh3ga@virginia.edu"),
    ("Eric Moore", "zqu3wh@virginia.edu"),
    ("Richard Charles Moreno", "jxm8cj@virginia.edu"),
    ("Sam John Palmer", "vwz2at@virginia.edu"),
    ("Charles Perry", "ghu7yg@virginia.edu"),
    ("Pierce Conor Seigne", "xrk9rs@virginia.edu"),
    ("Jonathan Logan Seyfert", "mea6wq@virginia.edu"),
    ("William Hayden Sheets", "fnh4mv@virginia.edu"),
    ("Alexander J Valencia", "nrk7dj@virginia.edu"),
    ("Jimmy Wischusen", "kna3ed@virginia.edu"),
]
DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

wb = Workbook()

# ============================================================ Week Plan
ws = wb.active
ws.title = "Week Plan"
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 50
ws.column_dimensions["F"].width = 44
ws.row_dimensions[1].height = 20

ws["A1"] = "HoosLog — Weekly Plan";  ws["A1"].font = title_f
ws["A3"] = "Week of (Monday):";      ws["A3"].font = label_f
ws["B3"] = dt.datetime(2026, 9, 7)
ws["B3"].font, ws["B3"].fill, ws["B3"].border = body_f, yellow_fill, bb
ws["B3"].number_format = "mm/dd/yyyy"
ws["C3"] = "← edit this one date; every day below updates itself"
ws["C3"].font = hint_f

for col, text in zip("ABCD", ["DAY", "DATE", "DISTANCE PLAN", "MID-DISTANCE PLAN"]):
    c = ws[f"{col}5"]
    c.value, c.font, c.fill, c.border = text, head_f, navy_fill, bb

for i, day in enumerate(DAYS):
    r = 6 + i
    ws.row_dimensions[r].height = 30
    a = ws.cell(row=r, column=1, value=day)
    a.font, a.border, a.alignment = label_f, bb, Alignment(vertical="top")
    b = ws.cell(row=r, column=2, value=f"=$B$3+{i}" if i else "=$B$3")
    b.font, b.border, b.number_format = body_f, bb, "ddd\\ m/d"
    b.alignment = Alignment(vertical="top")
    for col in (3, 4):
        p = ws.cell(row=r, column=col)
        p.font, p.fill, p.border, p.alignment = body_f, yellow_fill, bb, wrap_top

ws["F5"] = "HOW TO FILL THIS IN";           ws["F5"].font = label_f
for r, txt in [
    (6,  "Two columns = two schedules. Column C is what the"),
    (7,  "distance guys run, column D is what the mid-D guys run."),
    (8,  "Format is the same as the paper sheet, e.g."),
    (9,  '   "TR + hurdles + drills, WEIGHTS"'),
    (10, '   "3x1mi @ threshold, 2min rest"'),
    (11, "Leave a day blank for no plan."),
    (12, "Who is mid-D is set on the Goals tab, not here."),
]:
    ws[f"F{r}"] = txt; ws[f"F{r}"].font = hint_f

ws["A14"] = ("Yellow cells are yours: the Monday date, and each day's plan in BOTH columns. "
             "Leave a day blank for no plan. Don't add or delete rows or columns.")
ws["A14"].font = hint_f

# ============================================================ Goals
gs = wb.create_sheet("Goals")
gs.column_dimensions["A"].width = 26
gs.column_dimensions["B"].width = 30
gs.column_dimensions["C"].width = 20
gs.column_dimensions["D"].width = 16
gs.column_dimensions["F"].width = 52
gs.row_dimensions[1].height = 39
gs.freeze_panes = "A2"

for col, text in zip("ABCD", ["ATHLETE NAME", "UVA EMAIL", "WEEKLY GOAL (MILES)", "GROUP"]):
    c = gs[f"{col}1"]
    c.value, c.font, c.fill, c.border, c.alignment = text, head_f, navy_fill, bb, wrap_top

last = 1 + len(ROSTER)
for i, (name, email) in enumerate(ROSTER):
    r = 2 + i
    for col, val in ((1, name), (2, email), (3, None), (4, None)):
        c = gs.cell(row=r, column=col, value=val)
        c.font, c.border = body_f, bb
        if col == 3:                       # mileage — right-aligned so "62" and
            c.fill = yellow_fill           # "65-70" line up as one column
            c.alignment = Alignment(horizontal="right")
        elif col == 4:                     # group — centered, so it reads apart
            c.fill = yellow_fill           # from the mileage beside it
            c.alignment = Alignment(horizontal="center")

dv = DataValidation(type="list", formula1='"Distance,Mid-D"', allow_blank=True,
                    showDropDown=False, promptTitle="Training group",
                    prompt="Distance or Mid-D. Leave blank to keep the athlete where they already are.")
dv.error = "Type Distance or Mid-D, or leave it blank."
dv.errorTitle = "Distance or Mid-D"
gs.add_data_validation(dv)
dv.add(f"D2:D{last}")

for r, txt in [
    (1, "Yellow cells are yours: each athlete's weekly mileage, and their group."),
    (2, "MILEAGE — everyone gets a number, both groups. Plain (58), a range"),
    (3, '           (55-60), or a minimum (60+). Blank = no goal this week.'),
    (4, "GROUP — Distance or Mid-D. This is what decides which of the two"),
    (5, "           plan columns the athlete sees. Set it once; it sticks week"),
    (6, "           to week. Leave blank to leave an athlete where they are."),
    (7, "Names and emails are the current roster — don't retype them. The email"),
    (8, "is how the upload finds each athlete's account."),
]:
    gs[f"F{r}"] = txt; gs[f"F{r}"].font = hint_f
gs["F1"].font = Font(name="Arial", size=9, bold=True, color=NAVY)

# Sanity counters live in F/G on purpose. The importer walks Goals rows 2..N
# reading columns A-C; anything it finds there is treated as an athlete, so a
# label in column A would come back as "row 33 has no email" and refuse the
# whole upload. Columns E+ are outside the parsed range.
gs.column_dimensions["G"].width = 10
for off, (lbl, grp) in enumerate([("Mid-D:", "Mid-D"), ("Distance:", "Distance")]):
    r = last + 2 + off
    gs[f"F{r}"] = lbl; gs[f"F{r}"].font = label_f
    gs[f"G{r}"] = f'=COUNTIF($D$2:$D${last},"{grp}")'; gs[f"G{r}"].font = body_f
gs[f"F{last+4}"] = "Blank group cells aren't counted — those guys stay where they were."
gs[f"F{last+4}"].font = hint_f

# ============================================================ READ ME
rs = wb.create_sheet("READ ME")
rs.column_dimensions["A"].width = 104
rs.row_dimensions[1].height = 20
rs["A1"] = "HoosLog week-plan template — how it works"; rs["A1"].font = title_f
lines = [
    (3,  "1. Week Plan tab: set the Monday date (one yellow cell). Type the distance guys' workout in", body_f),
    (4,  "   column C and the mid-D guys' workout in column D. Leave a day blank for no plan.", body_f),
    (5,  "2. Goals tab: every athlete's weekly mileage, and their group (Distance or Mid-D).", body_f),
    (6,  "   Everyone gets mileage. The group only decides which workout column they see.", body_f),
    (7,  "3. Save, then drag this file into HoosLog (Coach → Post a week). You'll see exactly what it", body_f),
    (8,  "   says — including who's moving groups — before anything posts.", body_f),
    (10, "Group assignment sticks. Once a guy is marked Mid-D he stays Mid-D every week until you", hint_f),
    (11, "change that cell. A blank group cell changes nothing — it does not reset him to distance.", hint_f),
    (13, "Rules the upload depends on: don't rename tabs, don't add or delete rows or columns on the", hint_f),
    (14, "Week Plan tab, and emails on the Goals tab must match athletes' HoosLog login emails.", hint_f),
    (16, "Recreated from the team's paper weekly sheet — same structure, same workflow. (docs/mockups/10)", hint_f),
]
for r, txt, f in lines:
    rs[f"A{r}"] = txt; rs[f"A{r}"].font = f

out = "docs/templates/hooslog_week_plan_template.xlsx"
wb.save(out)
print("wrote", out, "| roster rows:", len(ROSTER), "| last goal row:", last)
