import openpyxl, datetime as dt, shutil, os
SRC = "docs/templates/hooslog_week_plan_template.xlsx"
OUT = os.path.expanduser("~/fixtures"); os.makedirs(OUT, exist_ok=True)
ROSTER = None

def load():
    return openpyxl.load_workbook(SRC)

def fill(wb, monday=dt.datetime(2026,9,7), dist=None, mid=None, rows=None):
    ws = wb["Week Plan"]; ws["B3"] = monday
    for i in range(7):
        ws.cell(row=6+i, column=3, value=(dist or [None]*7)[i])
        ws.cell(row=6+i, column=4, value=(mid or [None]*7)[i])
    gs = wb["Goals"]
    # clear roster mileage/group first
    for r in range(2, 32):
        gs.cell(row=r, column=3, value=None); gs.cell(row=r, column=4, value=None)
    for r, goal, grp in (rows or []):
        gs.cell(row=r, column=3, value=goal); gs.cell(row=r, column=4, value=grp)
    return wb

D = ["TR + drills","6x1mi @ threshold","TR 8-10mi","TR + strides","Fartlek","LR 14-16mi","Off"]
M = ["TR + drills","8x400 @ 3k","TR 6-8mi","TR + strides","6x200","LR 10-12mi","Off"]

# 1 both schedules + groups
fill(load(), dist=D, mid=M, rows=[(2,62,"Distance"),(3,45,"Mid-D"),(4,"55-60","Mid-D"),(5,"60+",None)]).save(f"{OUT}/01_normal.xlsx")
# 2 bad group word
fill(load(), dist=D, mid=M, rows=[(2,62,"Middle-ish")]).save(f"{OUT}/02_bad_group.xlsx")
# 3 mid athletes, empty mid column
fill(load(), dist=D, mid=None, rows=[(2,62,"Distance"),(3,45,"Mid-D")]).save(f"{OUT}/03_mid_no_plan.xlsx")
# 4 mid plan, nobody mid
fill(load(), dist=D, mid=M, rows=[(2,62,"Distance")]).save(f"{OUT}/04_plan_no_mid.xlsx")
# 5 group but no mileage
fill(load(), dist=D, mid=M, rows=[(2,None,"Mid-D")]).save(f"{OUT}/05_group_no_goal.xlsx")
# 6 distance column empty
fill(load(), dist=None, mid=M, rows=[(2,62,"Mid-D")]).save(f"{OUT}/06_no_distance.xlsx")
# 7 group spelling variants
fill(load(), dist=D, mid=M, rows=[(2,60,"md"),(3,60,"MID DISTANCE"),(4,60,"d"),(5,60,"  Mid-D ")]).save(f"{OUT}/07_variants.xlsx")
# 8 not a monday
fill(load(), monday=dt.datetime(2026,9,8), dist=D, mid=M, rows=[(2,62,"Distance")]).save(f"{OUT}/08_not_monday.xlsx")
print("fixtures written")
